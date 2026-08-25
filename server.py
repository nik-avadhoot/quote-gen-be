"""
CFB Quotation Master — Export Server  (v2.0)
============================================
Stateless Flask API. Its only job is to fill the Excel master template
(CFB_Quotation_Master_v7.xlsx) with quote data posted by the frontend and
return the workbook as a download.

Local development:
    Terminal 1:  python server.py          →  http://localhost:3001
    Terminal 2:  npm run dev               →  http://localhost:5173

Required packages (run once):
    pip install -r requirements.txt

Environment:
    CORS_ORIGINS  Comma-separated list of allowed browser origins.
                  Defaults to the local Vite dev server when unset.

Template:
    CFB_Quotation_Master_v7.xlsx must sit in the SAME folder as this file.
"""

# ═══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# All libraries this server needs. If any are missing, run:
#   pip install -r requirements.txt
# ═══════════════════════════════════════════════════════════════════════════════
import os
import io
import secrets
from datetime import datetime

from flask import Flask, request, send_file, jsonify, g
from flask_cors import CORS
import openpyxl

from supabase_client import get_supabase, get_supabase_admin
from auth import require_auth, require_role


# ═══════════════════════════════════════════════════════════════════════════════
# APP SETUP
# Must be defined BEFORE any @app.route decorators
# ═══════════════════════════════════════════════════════════════════════════════
app = Flask(__name__)

# Allowed browser origins. In production set CORS_ORIGINS to the deployed
# frontend URL(s), comma-separated. Falls back to the local Vite dev server.
DEFAULT_ORIGINS = ",".join([
    "https://quote-gen-fe.vercel.app",   # production frontend
    "http://localhost:5173",             # Vite dev server
    "http://127.0.0.1:5173",
])
CORS_ORIGINS = [
    o.strip()
    for o in os.environ.get("CORS_ORIGINS", DEFAULT_ORIGINS).split(",")
    if o.strip()
]
CORS(app, origins=CORS_ORIGINS, allow_headers=["Content-Type", "Authorization"])

# Path to the Excel master template — must sit beside this file
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "CFB_Quotation_Master_v7.xlsx")


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def num(v, default=0):
    """
    Safe numeric conversion.
    Returns `default` only when the value is truly missing (None or "").
    Returns 0 if the value exists but is not a valid number.
    """
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def set_val(ws, addr, val):
    """Set a worksheet cell value, preserving its existing style."""
    ws[addr].value = val


def clear_row(ws, r, cols):
    """Blank out all data-input columns in a given row."""
    for col in cols:
        ws[f"{col}{r}"].value = None


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE: /health  — quick check that the server and template are ready
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/health", methods=["GET"])
def health():
    """
    Frontend calls this on load to confirm the server is running, the Excel
    template file is present, and Supabase is reachable.
    Returns: { ok: true, template: true/false, supabase: true/false }
    """
    try:
        get_supabase()
        supabase_ok = True
    except Exception:
        supabase_ok = False

    return jsonify({
        "ok":       True,
        "template": os.path.exists(TEMPLATE_PATH),
        "path":     TEMPLATE_PATH,
        "supabase": supabase_ok,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTE: /export  — fill the Excel master template and send it to the browser
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/export", methods=["POST"])
@require_auth
def export_xlsx():
    """
    Receives the complete quote data from the frontend, fills the Excel
    master template (CFB_Quotation_Master_v7.xlsx), and returns the file
    as a download.
    """
    if not os.path.exists(TEMPLATE_PATH):
        return jsonify({"error": f"Template not found: {TEMPLATE_PATH}"}), 404

    data       = request.get_json(force=True)
    items      = data.get("items",   [])
    rates      = data.get("rates",   [])
    freight    = data.get("freight", {})
    fname      = data.get("filename", "CFB_Quote.xlsx")
    # Fix 9: read meta fields sent by the frontend
    quote_ref       = data.get("quoteRef",      "")
    maker_name      = data.get("makerName",     "")
    quote_date_str  = data.get("quoteDate",     "")
    effective_from  = data.get("effectiveFrom", "")
    effective_to    = data.get("effectiveTo",   "")

    wb     = openpyxl.load_workbook(TEMPLATE_PATH)
    ws_cbb = wb["CBB+PP"]
    ws_rm  = wb["RATE MASTER"]
    ws_def = wb["DEFAULTS"]

    # ── Update RATE MASTER ────────────────────────────────────────────────────
    for row in range(7, 30):
        code_cell = ws_rm.cell(row, 1)
        if not code_cell.value:
            continue
        code     = str(code_cell.value).strip()
        app_rate = next((r for r in rates if r.get("code") == code), None)
        if app_rate:
            ws_rm.cell(row, 3).value = num(app_rate.get("price"))
            ws_rm.cell(row, 5).value = num(app_rate.get("disc"),    1.5)
            ws_rm.cell(row, 6).value = num(app_rate.get("freight"), 0)

    # ── Update DEFAULTS freight matrix ────────────────────────────────────────
    # A4 (ruling): template reserves S14:S18 for custom delivery locations (5 rows).
    # The VLOOKUP in BK7 covers S4:S18. Rows below S18 are outside the lookup range
    # and would cause BK7 to return #N/A → IFERROR → 0.
    # Strategy:
    #   1. Update rates for locations already present in S4:S13 (standard locations).
    #   2. Write new locations (present in freight dict, absent from column S) into
    #      S14:S18 only. Raise an explicit server error if more than 5 new locations
    #      would be needed — do not silently drop or overwrite template structure.
    PLANT_COL     = {"Kolkata": 20, "Nagpur": 21, "Pune": 22}  # T=20, U=21, V=22
    LOC_COL       = 19   # Column S
    STD_ROW_RANGE = range(4, 14)   # S4:S13 — pre-printed standard locations
    CUSTOM_ROWS   = list(range(14, 19))  # S14:S18 — 5 reserved custom slots

    # Collect every delivery location the in-app freight state knows about
    all_locs = set()
    for plant_dict in freight.values():
        all_locs.update(plant_dict.keys())

    # Read existing location → row mapping from the template (S4:S18)
    existing_loc_rows = {}
    for row_idx in range(4, 19):
        cell_val = ws_def.cell(row_idx, LOC_COL).value
        if cell_val:
            existing_loc_rows[str(cell_val).strip()] = row_idx

    # Identify new locations not yet in column S
    new_locs = sorted(loc for loc in all_locs if loc not in existing_loc_rows)

    # Find which custom rows (S14:S18) are already occupied
    used_custom = {r for r in CUSTOM_ROWS if ws_def.cell(r, LOC_COL).value}
    free_custom  = [r for r in CUSTOM_ROWS if r not in used_custom]

    if len(new_locs) > len(free_custom):
        # Explicit error rather than silent drop — operator must expand the template
        overflow = new_locs[len(free_custom):]
        return jsonify({"error":
            f"Template freight-matrix overflow: {len(new_locs)} new locations need "
            f"{len(new_locs) - len(free_custom)} more reserved rows than S14:S18 provides. "
            f"Cannot write: {overflow}. Expand the DEFAULTS sheet and update PLANT_COL, "
            f"then re-export."}), 422

    # Assign new locations to free custom rows
    for loc, row_idx in zip(new_locs, free_custom):
        ws_def.cell(row_idx, LOC_COL).value = loc
        existing_loc_rows[loc] = row_idx

    # Write freight rates for every known location
    for loc, row_idx in existing_loc_rows.items():
        for plant, col_idx in PLANT_COL.items():
            plant_freight = freight.get(plant, {})
            if loc in plant_freight:
                ws_def.cell(row_idx, col_idx).value = num(plant_freight[loc])

    # ── CBB+PP header ─────────────────────────────────────────────────────────
    f0 = items[0]["spec"] if items else {}

    ws_cbb["D2"] = f0.get("client",   "")
    ws_cbb["B2"] = f0.get("delivery", "")
    ws_cbb["B3"] = f0.get("plant",    "Nagpur")
    ws_cbb["D3"] = f0.get("sector",   "")
    # Fix 9: write Quote Date (B4) from meta, fallback to today; Quote Ref to D4
    try:
        ws_cbb["B4"] = datetime.strptime(quote_date_str, "%Y-%m-%d") if quote_date_str else datetime.now()
    except ValueError:
        ws_cbb["B4"] = datetime.now()
    mat_codes = ", ".join(i["spec"].get("material_code", "") for i in items if i["spec"].get("material_code"))
    ws_cbb["D4"] = (f"{quote_ref} | {mat_codes}") if quote_ref else mat_codes

    # Rate parameters
    interest = num(f0.get("interest"),   0.5)
    margin   = num(f0.get("margin"),     8)
    waste    = num(f0.get("waste"),      5)
    conv_box = num(f0.get("convRate"),   7)
    waste_pp = num(f0.get("wastePP"),    5)
    conv_pp  = num(f0.get("convRatePP"), 12.5)  # A1-04: consistent with engine and INIT_SPEC default

    ws_cbb["BA3"] = conv_box
    ws_cbb["BA4"] = conv_pp
    ws_cbb["AY3"] = waste    / 100
    ws_cbb["AY4"] = waste_pp / 100
    ws_cbb["BJ3"] = interest / 100
    ws_cbb["BJ4"] = interest / 100

    # Freight override (blank = use VLOOKUP matrix)
    freight_override = f0.get("freightOverride", "")
    ws_cbb["BK4"] = (num(freight_override)
                     if freight_override not in ("", None) else None)

    # Box vs PP margin
    margin_pp = num(data.get("marginPP", margin), margin)
    ws_cbb["BM3"] = margin    / 100
    ws_cbb["BM4"] = margin_pp / 100

    # Add-ons defaults (row 3 = Box, row 4 = Board/PP)
    rs4 = next(
        (i["spec"] for i in items
         if not i["spec"].get("rowType") or i["spec"].get("rowType") == "Box"),
        f0
    )
    for col, key in [
        ("BB", "printing"), ("BC", "stitching"), ("BD", "coating"),
        ("BE", "handling"), ("BF", "moqCharge"), ("BG", "packing"),
        ("BH", "other"),    ("BI", "unloading"),
    ]:
        ws_cbb[f"{col}3"] = num(rs4.get(key, 0))
        ws_cbb[f"{col}4"] = 0   # Board/PP default

    # ── Data rows 7+ ──────────────────────────────────────────────────────────
    # B5-SERVER-04: Named constants for template row range.
    # DATA_START_ROW: first data row in the CBB+PP sheet (rows 1-6 are headers).
    # DATA_MAX_ROWS: maximum SKU rows the v7 template supports before stale data risk.
    DATA_START_ROW = 7
    DATA_MAX_ROWS  = 44   # template clears rows 7–50 (44 data rows); row 51+ untouched
    DATA_END_ROW   = DATA_START_ROW + DATA_MAX_ROWS   # 51 — first row NOT cleared

    if len(items) > DATA_MAX_ROWS:
        print(f"  ⚠  WARNING: Batch has {len(items)} items but template capacity is "
              f"{DATA_MAX_ROWS} rows. Rows beyond row {DATA_END_ROW - 1} will NOT be "
              f"cleared and may contain stale data from a previous export. "
              f"Consider splitting the quote into multiple exports.")

    DATA_COLS = [
        "B","C","D","E","F","G","H","I","J","K",
        "S","U","W","X","Y","Z",
        "BR","BS",                          # BR = SET Code, BS = Nos/Set (verified v7)
        # NOTE: column T = Calc GSM formula — never clear or overwrite it
        "AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM",
        "BB","BC","BD","BE","BF","BG","BH","BI",
    ]

    for idx, item in enumerate(items):
        r        = DATA_START_ROW + idx
        s        = item["spec"]
        row_type = s.get("rowType", "Box") or "Box"
        is_rs4   = row_type == "Box"

        ws_cbb[f"B{r}"] = row_type
        ws_cbb[f"C{r}"] = s.get("material_code", "")
        ws_cbb[f"D{r}"] = s.get("product", "")
        ws_cbb[f"E{r}"] = s.get("delivery", f0.get("delivery", ""))
        ws_cbb[f"I{r}"] = s.get("boxType", "RSC")
        ws_cbb[f"J{r}"] = int(num(s.get("ply", 5)))
        ws_cbb[f"K{r}"] = int(num(s.get("ups", 1)))

        # Issue 3 fix: write effective L/W/H for ALL row types directly.
        # For PP rows, s["L"] and s["W"] already hold resolved effective values
        # (set by autoCalcPPDims in sendAllToQuoteItems). H blank for flat pieces.
        ws_cbb[f"F{r}"] = num(s.get("L")) or None
        ws_cbb[f"G{r}"] = num(s.get("W")) or None
        ws_cbb[f"H{r}"] = None if not is_rs4 else (num(s.get("H")) or None)

        # Nos/Set — column BS (verified v7). Required for SET rate accumulation in Excel.
        ws_cbb[f"BS{r}"] = int(num(s.get("qtyPerSet"), 1) or 1)
        # SET Code — column BR.
        ws_cbb[f"BR{r}"] = s.get("setCode", "") or None
        ws_cbb[f"S{r}"]  = num(s.get("board_gsm")) or None
        ws_cbb[f"U{r}"]  = num(s.get("spec_bs"))   or None
        ws_cbb[f"W{r}"]  = num(s.get("spec_bct"))  or None
        ws_cbb[f"X{r}"]  = num(s.get("spec_ect"))  or None
        ws_cbb[f"Y{r}"]  = num(s.get("spec_cobb")) or None
        ws_cbb[f"Z{r}"]  = num(s.get("reqBoxWt"))  or None
        ws_cbb[f"AB{r}"] = s.get("flute_F1") or None
        ws_cbb[f"AC{r}"] = s.get("flute_F2") or None

        layers = s.get("layers", {})
        for cell_col, layer_key, sub_key in [
            ("AD", "TOP", "code"), ("AE", "TOP", "gsm"),
            ("AF", "F1",  "code"), ("AG", "F1",  "gsm"),
            ("AH", "L1",  "code"), ("AI", "L1",  "gsm"),
            ("AJ", "F2",  "code"), ("AK", "F2",  "gsm"),
            ("AL", "L2",  "code"), ("AM", "L2",  "gsm"),
        ]:
            raw = layers.get(layer_key, {}).get(sub_key, "")
            if sub_key == "gsm":
                ws_cbb[f"{cell_col}{r}"] = num(raw) or None
            else:
                ws_cbb[f"{cell_col}{r}"] = str(raw) if raw else None

        ws_cbb[f"BB{r}"] = num(s.get("printing",  0))
        ws_cbb[f"BC{r}"] = num(s.get("stitching", 0))
        ws_cbb[f"BD{r}"] = num(s.get("coating",   0))
        ws_cbb[f"BE{r}"] = num(s.get("handling",  0))
        ws_cbb[f"BF{r}"] = num(s.get("moqCharge", 0))
        ws_cbb[f"BG{r}"] = num(s.get("packing",   0))
        ws_cbb[f"BH{r}"] = num(s.get("other",     0))
        ws_cbb[f"BI{r}"] = num(s.get("unloading", 0))

        is_pp_row      = row_type in ("Plate", "Part-L", "Part-W")
        default_margin = margin_pp if is_pp_row else margin
        item_margin    = num(s.get("margin"), default_margin)

        if abs(item_margin - default_margin) > 0.001:
            ws_cbb[f"BM{r}"] = item_margin / 100
        elif is_pp_row:
            ws_cbb[f"BM{r}"] = "=$BM$4"

    # ── Clear remaining sample rows ───────────────────────────────────────────
    for r in range(DATA_START_ROW + len(items), DATA_END_ROW):
        for col in DATA_COLS:
            cell = ws_cbb[f"{col}{r}"]
            if cell.value not in (None, 0, ""):
                cell.value = None

    # ── Save and return ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        download_name=fname,
        as_attachment=True,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES: /auth/*  — login / refresh / logout for the frontend.
# The frontend never talks to Supabase directly; these are its only path to
# an authenticated session.
# ═══════════════════════════════════════════════════════════════════════════════

def _fetch_profile_or_none(user_id):
    try:
        resp = (
            get_supabase_admin()
            .table("profiles")
            .select("*")
            .eq("id", user_id)
            .single()
            .execute()
        )
        return resp.data
    except Exception:
        return None


@app.route("/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(force=True) or {}
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email and password are required"}), 400

    # Build the client OUTSIDE the credential try/except below. get_supabase()
    # raises RuntimeError when SUPABASE_URL / SUPABASE_PUBLISHABLE_KEY are
    # missing; caught alongside a genuine auth failure it surfaced as "Invalid
    # email or password", which sends you looking at the user record instead of
    # the deployment. A config fault must never be reportable as bad credentials.
    try:
        supabase = get_supabase()
    except RuntimeError as exc:
        app.logger.error("Supabase client unavailable: %s", exc)
        return jsonify({"error": "Auth backend is not configured"}), 500

    try:
        auth_resp = supabase.auth.sign_in_with_password(
            {"email": email, "password": password}
        )
    except Exception:
        return jsonify({"error": "Invalid email or password"}), 401

    session, user = auth_resp.session, auth_resp.user
    if not session or not user:
        return jsonify({"error": "Invalid email or password"}), 401

    profile = _fetch_profile_or_none(user.id)
    if not profile:
        return jsonify({"error": "No profile found for this account"}), 401
    if not profile.get("active", False):
        return jsonify({"error": "Account is deactivated"}), 403

    return jsonify({
        "access_token": session.access_token,
        "refresh_token": session.refresh_token,
        "expires_at": session.expires_at,
        "profile": {**profile, "email": user.email},
    })


@app.route("/auth/refresh", methods=["POST"])
def auth_refresh():
    data = request.get_json(force=True) or {}
    refresh_token = data.get("refresh_token")
    if not refresh_token:
        return jsonify({"error": "refresh_token is required"}), 400

    try:
        auth_resp = get_supabase().auth.refresh_session(refresh_token)
    except Exception:
        return jsonify({"error": "Invalid or expired refresh token"}), 401

    session, user = auth_resp.session, auth_resp.user
    if not session or not user:
        return jsonify({"error": "Invalid or expired refresh token"}), 401

    profile = _fetch_profile_or_none(user.id)
    if not profile:
        return jsonify({"error": "No profile found for this account"}), 401
    if not profile.get("active", False):
        return jsonify({"error": "Account is deactivated"}), 403

    return jsonify({
        "access_token": session.access_token,
        "refresh_token": session.refresh_token,
        "expires_at": session.expires_at,
        "profile": {**profile, "email": user.email},
    })


@app.route("/auth/logout", methods=["POST"])
@require_auth
def auth_logout():
    # Best-effort — revoke every refresh token for this session. The frontend
    # clears its own stored tokens regardless of whether this succeeds.
    try:
        get_supabase_admin().auth.admin.sign_out(g.access_token, "global")
    except Exception:
        pass
    return jsonify({"ok": True})


@app.route("/auth/me", methods=["PATCH"])
@require_auth
def auth_update_me():
    """
    Self-service profile edit — a user updating their own display_name/plant.
    Deliberately does not accept role or active: those stay admin-only via
    /admin/users/<uid>, even for a user editing their own row.
    """
    data = request.get_json(force=True) or {}
    updates = {}
    if "display_name" in data:
        display_name = (data["display_name"] or "").strip()
        if not display_name:
            return jsonify({"error": "Display name cannot be empty"}), 400
        updates["display_name"] = display_name
    if "plant" in data:
        updates["plant"] = (data["plant"] or "").strip() or None

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    result = (
        get_supabase_admin()
        .table("profiles")
        .update(updates)
        .eq("id", g.current_user["id"])
        .execute()
    )
    if not result.data:
        return jsonify({"error": "Could not update profile"}), 400
    return jsonify({**result.data[0], "email": g.current_user["email"]})


@app.route("/auth/change-password", methods=["POST"])
@require_auth
def auth_change_password():
    """
    Self-service password change — any logged-in user, for their own account.
    Requires the current password (re-verified via sign_in_with_password) so
    a leftover valid access token on a shared machine can't be used to lock
    the real owner out.
    """
    data             = request.get_json(force=True) or {}
    current_password = data.get("current_password") or ""
    new_password     = data.get("new_password") or ""

    if not current_password or not new_password:
        return jsonify({"error": "current_password and new_password are required"}), 400
    if len(new_password) < 8:
        return jsonify({"error": "New password must be at least 8 characters"}), 400

    try:
        get_supabase().auth.sign_in_with_password({
            "email": g.current_user["email"],
            "password": current_password,
        })
    except Exception:
        return jsonify({"error": "Current password is incorrect"}), 401

    try:
        get_supabase_admin().auth.admin.update_user_by_id(
            g.current_user["id"], {"password": new_password}
        )
    except Exception as e:
        return jsonify({"error": f"Could not update password: {e}"}), 400

    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES: /admin/users*  — Admin-only user management.
# ═══════════════════════════════════════════════════════════════════════════════

VALID_ROLES = ("maker", "checker", "admin")


@app.route("/admin/users", methods=["GET"])
@require_auth
@require_role("admin")
def list_users():
    profiles_resp = get_supabase_admin().table("profiles").select("*").execute()
    profiles_by_id = {p["id"]: p for p in profiles_resp.data}

    auth_users = get_supabase_admin().auth.admin.list_users()
    result = []
    for u in auth_users:
        profile = profiles_by_id.get(u.id)
        if not profile:
            continue
        result.append({
            **profile,
            "email": u.email,
            "last_sign_in_at": u.last_sign_in_at.isoformat() if u.last_sign_in_at else None,
        })
    return jsonify({"users": result})


@app.route("/admin/users", methods=["POST"])
@require_auth
@require_role("admin")
def create_user():
    data = request.get_json(force=True) or {}
    email        = (data.get("email") or "").strip()
    display_name = (data.get("display_name") or "").strip()
    role         = data.get("role", "maker")
    plant        = data.get("plant") or None

    if not email or not display_name:
        return jsonify({"error": "email and display_name are required"}), 400
    if role not in VALID_ROLES:
        return jsonify({"error": f"role must be one of {VALID_ROLES}"}), 400

    generated = not data.get("password")
    password  = data.get("password") or secrets.token_urlsafe(9)

    try:
        created = get_supabase_admin().auth.admin.create_user({
            "email": email,
            "password": password,
            "email_confirm": True,
        })
    except Exception as e:
        return jsonify({"error": f"Could not create auth user: {e}"}), 400

    uid = created.user.id
    try:
        get_supabase_admin().table("profiles").insert({
            "id": uid,
            "display_name": display_name,
            "role": role,
            "plant": plant,
            "active": True,
        }).execute()
    except Exception as e:
        # Don't leave an orphaned auth user with no profile row behind.
        get_supabase_admin().auth.admin.delete_user(uid)
        return jsonify({"error": f"Could not create profile: {e}"}), 400

    resp = {
        "id": uid, "email": email, "display_name": display_name,
        "role": role, "plant": plant, "active": True,
    }
    if generated:
        resp["temp_password"] = password
    return jsonify(resp), 201


@app.route("/admin/users/<uid>", methods=["PATCH"])
@require_auth
@require_role("admin")
def update_user(uid):
    data = request.get_json(force=True) or {}
    updates = {}

    if "display_name" in data:
        updates["display_name"] = data["display_name"]
    if "plant" in data:
        updates["plant"] = data["plant"]
    if "role" in data:
        if data["role"] not in VALID_ROLES:
            return jsonify({"error": f"role must be one of {VALID_ROLES}"}), 400
        if uid == g.current_user["id"] and data["role"] != "admin":
            return jsonify({"error": "You cannot change your own role"}), 400
        updates["role"] = data["role"]
    if "active" in data:
        if uid == g.current_user["id"] and not data["active"]:
            return jsonify({"error": "You cannot deactivate your own account"}), 400
        updates["active"] = bool(data["active"])

    if not updates:
        return jsonify({"error": "No fields to update"}), 400

    result = get_supabase_admin().table("profiles").update(updates).eq("id", uid).execute()
    if not result.data:
        return jsonify({"error": "User not found"}), 404
    return jsonify(result.data[0])


@app.route("/admin/users/<uid>/reset-password", methods=["POST"])
@require_auth
@require_role("admin")
def reset_password(uid):
    data = request.get_json(force=True) or {}
    generated = not data.get("password")
    password  = data.get("password") or secrets.token_urlsafe(9)

    try:
        get_supabase_admin().auth.admin.update_user_by_id(uid, {"password": password})
    except Exception as e:
        return jsonify({"error": f"Could not reset password: {e}"}), 400

    resp = {"ok": True}
    if generated:
        resp["temp_password"] = password
    return jsonify(resp)


# ═══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Local development entry point only. On Vercel the app is served as a
    # WSGI application via api/index.py and this block never runs.
    print("\n" + "=" * 55)
    print("  CFB Quotation Master — Export Server  v2.0")
    print("=" * 55)
    print(f"  Template       : {TEMPLATE_PATH}")
    print(f"  Template found : {os.path.exists(TEMPLATE_PATH)}")
    print(f"  CORS origins   : {', '.join(CORS_ORIGINS)}")
    print(f"  Listening on   : http://localhost:3001")
    print("=" * 55 + "\n")

    app.run(port=3001, debug=False)
